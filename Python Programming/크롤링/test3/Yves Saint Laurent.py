import requests
import urllib
from openpyxl import Workbook
from bs4 import BeautifulSoup

wb = Workbook()
ws = wb.active
ws.title = "first_Steet"
ws["A1"] = "brand"
ws["B1"] = "name"
ws["C1"] = "price"
ws["D1"] = "R"
ws["E1"] = "G"
ws["F1"] = "B"
ws["G1"] = "src"


def spider(max_pages):

    page = 1
    urlpage = 0
    i = 2

    while page <= max_pages:
        url1 = 'https://www.yslbeautykr.com/ko_KR/makeup/lips?sz=12&start='
        url = url1 + str(urlpage)
        source_code = requests.get(url)
        plain_text = source_code.text
        soup = BeautifulSoup(plain_text, 'html.parser')

        for div in soup.findAll('div', {'class': 'product_tile b-product_tile'}):

            p_tag = div.find('a', {'class': 'product_name mobile_content'}).getText()
            name_tag = div.find('div', {'class': 'product_subtitle'}).getText()
            span_tag = div.find('p', {'class': 'product_price price_sale b-product_price-sale b-product_price'}).getText()
            img_tag = div.find('img')

            print(p_tag.strip())
            print(name_tag.strip())
            print(span_tag.strip())
            print(img_tag.get('data-desktop-src').strip())

            nametemp = p_tag.strip()
            if len(nametemp) == 0:
                filename = str(i)
                i = i + 1
            else:
                filename = nametemp

            imagesfile = open(filename + ".jpeg", 'wb')
            imagesfile.write(urllib.request.urlopen(img_tag.get('data-desktop-src')).read())
            imagesfile.close()

            ws.cell(row=i, column=1, value=p_tag.strip())
            ws.cell(row=i, column=2, value=name_tag.strip())
            ws.cell(row=i, column=3, value=span_tag.strip())
            ws.cell(row=i, column=4, value='')
            ws.cell(row=i, column=5, value='')
            ws.cell(row=i, column=6, value='')
            ws.cell(row=i, column=7, value=p_tag.strip() + ".jpeg")
            # ws.cell(row=i, column=5, value=img_tag.get('data-desktop-src').strip())
            i = i + 1

        page += 1
        urlpage += 12


spider(1)

wb.save("Yves Saint Laurent.xlsx")
