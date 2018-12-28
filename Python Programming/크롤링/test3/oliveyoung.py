import requests
import urllib
from openpyxl import Workbook
from bs4 import BeautifulSoup

wb = Workbook()
ws = wb.active
ws.title = "first_Steet"
ws["A1"] = "brand"
ws["B1"] = "name"
ws["C1"] = "price"
ws["D1"] = "src"


def spider(max_pages):

    page = 1
    urlpage = 1
    i = 2

    while page <= max_pages:
        url1 = 'http://www.oliveyoung.co.kr/store/display/getMCategoryList.do?dispCatNo=100000100020006&fltDispCatNo=&prdSort=01&pageIdx='
        url2 = '&rowsPerPage=24&searchTypeSort=btn_thumb&plusButtonFlag=N'
        url = url1 + str(urlpage) + url2
        source_code = requests.get(url)
        plain_text = source_code.text
        soup = BeautifulSoup(plain_text, 'html.parser')

        for div in soup.findAll('div', {'class': 'prd_info'}):

            p_tag = div.find('span', {'class': 'tx_brand'}).getText()
            name_tag = div.find('p', {'class': 'tx_name'}).getText()
            span_tag = div.find('span', {'class': 'tx_num'}).getText()
            img_tag = div.find('img')

            print(p_tag)
            print(name_tag)
            print(span_tag)
            print(img_tag.get('src'))

            # nametemp = img_tag.get('alt')
            # if len(nametemp) == 0:
            #     filename = str(i)
            #     i = i + 1
            # else:
            #     filename = nametemp
            #
            # imagesfile = open(filename + ".jpeg", 'wb')
            # imagesfile.write(urllib.request.urlopen(img_tag.get('src')).read())
            # imagesfile.close()

            ws.cell(row=i, column=1, value=p_tag)
            ws.cell(row=i, column=2, value=name_tag)
            ws.cell(row=i, column=3, value=span_tag)
            ws.cell(row=i, column=4, value=img_tag.get('src'))
            i = i + 1

        page += 1
        urlpage += 1


spider(10)

wb.save("test1.xlsx")
